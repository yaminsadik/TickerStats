from flask import Blueprint, request, send_file, jsonify
import yfinance as yf
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import numpy as np
from openpyxl.styles import Font



export_xl_bp = Blueprint('export_xl', __name__)

@export_xl_bp.route('/api/export-xlsx')
def export_xlsx():
    tickers = request.args.get('ticker')
    if not tickers:
        return jsonify({"error": "Ticker is required"}), 400

    tickers = [t.strip().upper() for t in tickers.split(',')]
    records = []

    for ticker in tickers:
        try:
            stock = yf.Ticker(ticker)
            info = stock.info

            record = {
                "Ticker": ticker,
                "Company": info.get("longName"),
                "Share Price": round(info.get("currentPrice", 0), 2),
                "Market Cap (B)": round(info.get("marketCap", 0) / 1e9, 2),
                "Enterprise Value (B)": round(info.get("enterpriseValue", 0) / 1e9, 2),
                "Forward P/E": round(info.get("forwardPE", 0), 2),
                "Price/Sales": round(info.get("priceToSalesTrailing12Months", 0), 2),
                "Price/Book": round(info.get("priceToBook", 0), 2),
                "EV/EBITDA": round(info.get("enterpriseToEbitda", 0), 2),
                "EV/Revenue": round(info.get("enterpriseToRevenue", 0), 2),
                "Profit Margin (%)": f"{round(info.get('profitMargins', 0) * 100, 2)}%",
                "ROA (%)": f"{round(info.get('returnOnAssets', 0) * 100, 2)}%",
                "ROE (%)": f"{round(info.get('returnOnEquity', 0) * 100, 2)}%",
                "Debt/Equity (%)": f"{round(info.get('debtToEquity', 0), 2)}%",
                "Beta": round(info.get("beta", 0), 2)
            }
            records.append(record)
        except Exception:
            continue

    df = pd.DataFrame(records)

    # Excel with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Financials"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Header formatting
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    thresholds = {
        "Forward P/E": (23.79, "lower"),
        "Price/Sales": (3.69, "lower"),
        "Price/Book": (8.91, "lower"),
        "EV/EBITDA": (19.77, "lower"),
        "EV/Revenue": (4.09, "lower"),
        "Profit Margin (%)": (13.3, "higher"),
        "ROA (%)": (7.75, "higher"),
        "ROE (%)": (25.88, "higher"),
        "Debt/Equity (%)": (145.25, "lower"),
        "Beta": (1.13, "lower")
    }

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    headers = [cell.value for cell in ws[1]]
    num_data_rows = len(df)

    for col_index, col in enumerate(ws.iter_cols(min_row=2, max_row=1 + num_data_rows, min_col=1, max_col=ws.max_column)):
        metric = headers[col_index]
        if metric in thresholds:
            threshold, direction = thresholds[metric]
            for cell in col:
                try:
                    val = float(str(cell.value).replace('%', ''))
                    if direction == "lower":
                        cell.fill = green_fill if val <= threshold else red_fill
                    else:
                        cell.fill = green_fill if val >= threshold else red_fill
                except:
                    continue

    # Add summary rows
    ws.append([])
    summary_labels = ['Average', 'Minimum', 'Maximum', 'St. Dev']

    for label in summary_labels:
        row = [label, '']
        for col_index in range(2, len(headers)):
            values = []
            for i in range(2, 2 + num_data_rows):
                try:
                    val = ws.cell(row=i, column=col_index + 1).value
                    val = float(str(val).replace('%', ''))  # strip %
                    values.append(val)
                except:
                    continue
            if not values:
                row.append("N/A")
                continue

            if label == 'Average':
                val = np.mean(values)
            elif label == 'Minimum':
                val = np.min(values)
            elif label == 'Maximum':
                val = np.max(values)
            elif label == 'St. Dev':
                val = np.std(values)
            else:
                val = "N/A"

            header = headers[col_index]
            is_percent = header and header.endswith("(%)")
            row.append(f"{round(val, 2)}%" if is_percent else f"{round(val, 2)}")

        ws.append(row)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
        
    for row in ws.iter_rows(min_row=ws.max_row - 3, max_row=ws.max_row):
        row[0].font = Font(bold=True)


    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="financial_summary.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
