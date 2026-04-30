"""
Shared export helpers for CSV, XLSX, and PDF generation.
"""

import csv
import io
import logging
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)


def build_table_rows(
    rows_data: List[Dict[str, Any]],
    validated_fields: List[str],
    validated_perf_metrics: Optional[List[str]],
    include_dcf: bool,
) -> tuple[List[str], List[Dict[str, Any]]]:
    """
    Build normalised headers and flat row dicts from the raw service data.
    Returns (headers, flat_rows).
    """
    headers = ["symbol"] + list(validated_fields)
    if validated_perf_metrics:
        headers.extend(validated_perf_metrics)
    if include_dcf:
        headers.extend(["dcfPrice", "dcfUpside"])

    flat_rows: List[Dict[str, Any]] = []
    for row in rows_data:
        flat: Dict[str, Any] = {"symbol": row["symbol"]}
        for field in validated_fields:
            val = row["snapshot"].get(field)
            flat[field] = val if val is not None else ""
        if validated_perf_metrics and row.get("performance"):
            for metric in validated_perf_metrics:
                val = row["performance"].get(metric)
                flat[metric] = val if val is not None else ""
        elif validated_perf_metrics:
            for metric in validated_perf_metrics:
                flat[metric] = ""
        if include_dcf and row.get("dcf"):
            flat["dcfPrice"] = row["dcf"].get("dcfPrice", "")
            flat["dcfUpside"] = row["dcf"].get("dcfUpside", "")
        elif include_dcf:
            flat["dcfPrice"] = ""
            flat["dcfUpside"] = ""
        flat_rows.append(flat)

    return headers, flat_rows


def generate_csv(headers: List[str], flat_rows: List[Dict[str, Any]]) -> str:
    """Return CSV content as a string."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in flat_rows:
        writer.writerow(row)
    content = output.getvalue()
    output.close()
    return content


def generate_xlsx(headers: List[str], flat_rows: List[Dict[str, Any]]) -> bytes:
    """Return an XLSX workbook as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Relative Table"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, flat_row in enumerate(flat_rows, 2):
        for col_idx, header in enumerate(headers, 1):
            val = flat_row.get(header, "")
            # Try to convert numeric strings
            if isinstance(val, str) and val != "":
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-size columns
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row_idx in range(2, len(flat_rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_length = max(max_length, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 3, 30)

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_pdf(headers: List[str], flat_rows: List[Dict[str, Any]]) -> bytes:
    """Return a PDF document with the table as bytes."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("Relative Table Export", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Build table data
    table_data = [headers]
    for flat_row in flat_rows:
        row_vals = []
        for h in headers:
            val = flat_row.get(h, "")
            if isinstance(val, float):
                val = f"{val:.4f}" if abs(val) < 1 else f"{val:.2f}"
            row_vals.append(str(val) if val is not None else "")
        table_data.append(row_vals)

    # Calculate column widths based on content
    num_cols = len(headers)
    available_width = landscape(A4)[0] - 1.0 * inch
    col_width = available_width / num_cols

    table = Table(table_data, colWidths=[col_width] * num_cols)

    # Style the table
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ])
    table.setStyle(style)

    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()
